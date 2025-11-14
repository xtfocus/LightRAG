"""
Excel file processor.
 
This is a simple example processor for .xls, .xlsx, .xlsb files.
 
To create a new processor:
1. Copy this file and rename it (e.g., processing_docx.py)
2. Update the class name and supported_extensions
3. Implement the process_file method with your file type's extraction logic
4. Register the processor in the __init__.py file
"""

import ast
from functools import wraps
import os
from pathlib import Path
import logging
from logging import getLogger
import time
from typing import Tuple

import dotenv
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
import xlrd
from pyxlsb import open_workbook

from .base_processor import BaseFileProcessor, ProcessingResult
from lightrag.utils import logger

import requests

logger = getLogger(__name__)

root = Path(__file__).resolve().parent
env_path = root / ".env"
dotenv.load_dotenv(env_path)
LLM_API_BASE = os.getenv("LLM_API_BASE")
LLM_API_TOKEN = os.getenv("LLM_API_TOKEN")

def timer(logger=None):
    """
    Decorator function that logs the running time of a function when called.
    Args:
        logger (Optional[Logger]): The logger object to use for logging.
            If None, uses a default logger. Default is None.
    Returns:
        decorator: Decorator function.
    """
    if logger is None:
        logger = logging.getLogger(__name__)

    def decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            start_time = time.time()
            result = func(*args, **kwargs)
            end_time = time.time()
            elapsed_time = end_time - start_time
            logger.info("%s: running time: %.4f seconds",
                        func.__name__, elapsed_time)
            return result

        return wrapper

    return decorator


def async_timer(func):
    @wraps(func)
    async def wrapper(*args, **kwargs):
        start = time.perf_counter()
        result = await func(*args, **kwargs)
        end = time.perf_counter()
        print(f"Function {func.__name__} took {end - start:.4f} seconds.")
        return result
    return wrapper


class LLMClient:
    def __init__(self, api_base: str, token: str, model: str = "gpt-4.1"):
        self.token = token
        self.api_base = api_base
        self.model = model

    @async_timer
    async def detect_block(self, cells_data):
        prompt = self._build_prompt(cells_data)
        headers = {
            "Authorization": f"Bearer {self.token}",
            "Content-Type": "application/json",
        }
        payload = {
            "model": self.model,
            "messages": [
                {
                    "role": "system",
                    "content": "You are an expert at extracting context and structure from Excel data.",
                },
                {"role": "user", "content": prompt},
            ],
            "temperature": 0.1,
            "max_tokens": 5000,
        }
        # For async requests, use httpx or aiohttp, but here is sync example:
        response = requests.post(
            f"{self.api_base}/v1/chat/completions", headers=headers, json=payload
        )
        response.raise_for_status()
        content = response.json()["choices"][0]["message"]["content"]
        block_info = self._parse_blocks(content)
        return block_info

    def _build_prompt(self, cells_data):
        cell_lines = [f"{loc}: {val}" for loc, val in cells_data]
        cells_str = "\n".join(cell_lines)
        prompt = (
            "Given the following Excel cell data (location: value), "
            "identify blocks of text and tables. "
            "Return a list in Python format: [('text', location), ('table', start_location, end_location), ...].\n\n"
            f"Cells:\n{cells_str}\n\n"
            "Output only the Python list."
        )
        return prompt

    def _parse_blocks(self, response_text):
        try:
            return ast.literal_eval(response_text)
        except Exception:
            return []


class ExcelFileProcessor(BaseFileProcessor):
    """Processor for plain text (.txt) files.

    This processor handles UTF-8 encoded text files. It validates the content
    and ensures it's not empty or binary data.
    """

    def __init__(self):
        super().__init__()
        self.llm_client = LLMClient(
            api_base=LLM_API_BASE, token=LLM_API_TOKEN
        )

    @property
    def supported_extensions(self) -> tuple[str, ...]:
        """Return supported file extensions for this processor.

        Returns:
            tuple[str, ...]: Tuple of supported extensions (e.g., (".xlsx", ".xls", ".xlsb"))
        """
        return (
            ".xlsx",
            ".xls",
            ".xlsb",
        )

    @async_timer
    async def process_file(
        self,
        file_path: Path,
        file_bytes: bytes,
        file_size: int,
        track_id: str,
    ) -> ProcessingResult:
        """Process a .txt file and extract its text content.

        This method:
        1. Decodes the file bytes as UTF-8
        2. Validates the content is not empty
        3. Checks for binary data representation
        4. Returns the extracted text content

        Args:
            file_path: Path to the file being processed
            file_bytes: Raw file content as bytes
            file_size: Size of the file in bytes (for error reporting)
            track_id: Tracking ID for this processing operation

        Returns:
            ProcessingResult: Result containing:
                - success: True if processing succeeded
                - content: Extracted text content (empty if failed)
                - error_description: Human-readable error message (if failed)
                - original_error: Original exception message (if failed)
        """
        try:
            # Step 1: Decode the file as UTF-8
            # content = file_bytes.decode("utf-8")
            content = await self.extract_excel_markdown(file_path)

            # Step 2: Validate content is not empty
            if not content or len(content.strip()) == 0:
                return ProcessingResult(
                    success=False,
                    content="",
                    error_description="[File Extraction]Empty file content",
                    original_error="File contains no content or only whitespace",
                )

            # Step 3: Check for binary data representation
            # Sometimes binary data gets incorrectly decoded as text
            # if content.startswith("b'") or content.startswith('b"'):
            #     return ProcessingResult(
            #         success=False,
            #         content="",
            #         error_description="[File Extraction]Binary data in text file",
            #         original_error="File appears to contain binary data representation instead of text",
            #     )

            # Step 4: Success - return the content
            return ProcessingResult(
                success=True,
                content=content,
            )

        # except UnicodeDecodeError as e:
        #     # Handle encoding errors
        #     return ProcessingResult(
        #         success=False,
        #         content="",
        #         error_description="[File Extraction]UTF-8 encoding error, please convert it to UTF-8 before processing",
        #         original_error=f"File is not valid UTF-8 encoded text: {str(e)}",
        #     )

        except Exception as e:
            # Handle any other unexpected errors
            logger.error(
                f"[File Extraction]Unexpected error processing TXT file {file_path.name}: {str(e)}"
            )
            return ProcessingResult(
                success=False,
                content="",
                error_description="[File Extraction]TXT file processing error",
                original_error=f"Unexpected error: {str(e)}",
            )

    @async_timer
    async def extract_excel_markdown(self, file_path: Path | str) -> str:
        """
        Extract excel file of multiple types to markdown contents.
        Args:
            file_path (str): path to excel files
        Returns:
            contents (str): markdown content of the excel files
        """
        file_path = Path(file_path)
        assert file_path.exists(), "File not found error"
        assert file_path.suffix in [
            ".xlsx", ".xls", ".xlsb"], "Unsupport file type"
        if file_path.suffix == ".xlsx":
            contents = await self.extract_xlsx_markdown(file_path)
        elif file_path.suffix == ".xls":
            contents = await self.extract_xls_markdown(file_path)
        else:
            contents = await self.extract_xlsb_markdown(file_path)
        return contents

    @async_timer
    async def extract_xlsx_markdown(
        self, file_path: Path | str, limit: Tuple = (500, 250)
    ) -> str:
        """
        Extract xlsx file into markdown contents.
        Args:
            file_path (str): path to excel files.
            limit (Tuple): limit max row / max column
        Returns:
            contents (str): markdown content of the xlsx files.
        """
        file_name = Path(file_path).name
        markdown_chunks = [f"# Document: {file_name}"]

        data = {}
        wb = load_workbook(file_path, data_only=True)
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            sheet_content = []
            if sheet.max_row >= limit[0] or sheet.max_column >= limit[1]:
                sheet_content = await self.extract_large_xlsx(sheet)
            else:
                sheet_content = await self.extract_common_xlsx(sheet)
            data[sheet_name] = sheet_content

            markdown_chunks.append(f"\n## Sheet: {sheet_name}\n")
            markdown_chunks.append(sheet_content)

        # Combine all markdown chunks
        markdown_content = "\n".join(markdown_chunks)
        return markdown_content

    @async_timer
    async def extract_large_xlsx(self, sheet) -> str:
        """
        Converts the entire sheet to a markdown table.
        Args:
            sheet (openpyxl.Worksheet): The worksheet object.
        Returns:
            sheet_content (str): The markdown table as a string.
        """
        n_rows = sheet.max_row
        first_pct = 10 # max(1, int(n_rows * 0.05))

        # Read first 5% rows
        candidate_rows = []
        for i in range(1, first_pct + 1):
            row_values = [
                str(cell.value).replace("\n", "<br>") if cell.value is not None else "" for cell in sheet[i]]
            candidate_rows.append(row_values)

        # Find header row: most non-empty cells
        filled_counts = [
            sum(1 for cell in row if cell.strip()) for row in candidate_rows
        ]
        header_idx = filled_counts.index(max(filled_counts))

        header_row = candidate_rows[header_idx]
        # Free text rows above header
        markdown_lines = []
        for row in candidate_rows[:header_idx]:
            free_text = " ".join(cell for cell in row if cell.strip())
            if free_text:
                markdown_lines.append(f"{free_text}\n")

        # Markdown table header
        markdown_lines.append("| " + " | ".join(header_row) + " |")
        markdown_lines.append("|" + " --- |" * len(header_row))

        # Read rest of the sheet
        for i in range(header_idx + 2, n_rows + 1):  # +2 because openpyxl is 1-indexed
            row = [str(cell.value) if cell.value is not None else "" for cell in sheet[i]]
            markdown_lines.append("| " + " | ".join(row) + " |")

        sheet_content = "\n".join(markdown_lines)
        return sheet_content

    @staticmethod
    def process_cell_value(value: str):
        value = str(value).strip()
        value_list = value.split("\n")
        value_list = [line.replace("-", "\t", 1) if line.lstrip().startswith(
            "-") else line for line in value_list]
        value = "<br>".join(value_list)

        return value

    @async_timer
    async def extract_common_xlsx(self, sheet: Path | str) -> str:
        """
        Converts the sheet into a markdown data.
        Args:
            sheet (openpyxl.Worksheet): The worksheet object.
        Returns:
            sheet_content (str): The markdown data.
        """

        # get the list of cell values
        cells_data = []
        trunc_cells = []
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    col_letter = get_column_letter(cell.column)
                    value = self.process_cell_value(cell.value)
                    cells_data.append([f"{cell.row}{col_letter}", value])
                    trunc_cells.append([f"{cell.row}{col_letter}", value[:5]])

        # # Step 1: Call LLM to get the information of text and block with LLM
        block_info = await self.llm_client.detect_block(trunc_cells)

        # # Step 2: Process blocks
        markdown_chunks = []
        for block in block_info:
            if block[0] == "text":
                # Find and extract text for this cell
                loc = block[1]
                text_cell = next((v for l, v in cells_data if l == loc), "")
                if text_cell:
                    markdown_chunks.append(f"{text_cell}\n")
            elif block[0] == "table":
                start_loc, end_loc = block[1], block[2]
                start_row = int("".join(filter(str.isdigit, start_loc)))
                start_col = "".join(filter(str.isalpha, start_loc))
                end_row = int("".join(filter(str.isdigit, end_loc)))
                end_col = "".join(filter(str.isalpha, end_loc))
                # Convert column letter to index

                start_col_idx = column_index_from_string(start_col)
                end_col_idx = column_index_from_string(end_col)
                # Extract table block
                table_rows = []
                for r in range(start_row, end_row + 1):
                    row_cells = []
                    for c in range(start_col_idx, end_col_idx + 1):
                        cell = sheet.cell(row=r, column=c)
                        row_cells.append(
                            self.process_cell_value(
                                cell.value) if cell.value is not None else ""
                        )
                    table_rows.append(row_cells)
                # Markdown table
                if table_rows:
                    header = table_rows[0]
                    markdown_chunks.append("| " + " | ".join(header) + " |")
                    markdown_chunks.append("|" + " --- |" * len(header))
                    for row in table_rows[1:]:
                        markdown_chunks.append("| " + " | ".join(row) + " |")
        sheet_content = "\n".join(markdown_chunks)
        return sheet_content

    @async_timer
    async def extract_xls_markdown(
        self, file_path: Path | str, limit: Tuple = (500, 250)
    ) -> str:
        """
        Extract xls file into markdown contents.
        Args:
            file_path (str): path to excel files.
            limit (Tuple): limit max row / max column
        Returns:
            contents (str): markdown content of the xls files.
        """
        file_name = Path(file_path).name
        markdown_chunks = [f"# Document: {file_name}"]

        wb = xlrd.open_workbook(file_path)
        for sheet in wb.sheets():
            n_rows = sheet.nrows
            n_cols = sheet.ncols

            if n_rows >= limit[0] or n_cols >= limit[1]:
                sheet_content = await self.extract_large_xls(sheet)
            else:
                sheet_content = await self.extract_common_xls(sheet)

            markdown_chunks.append(f"\n## Sheet: {sheet.name}\n")
            markdown_chunks.append(sheet_content)

        markdown_content = "\n".join(markdown_chunks)
        return markdown_content

    @async_timer
    async def extract_large_xls(self, sheet) -> str:
        """
        Converts the entire xlrd sheet to a markdown table.
        Args:
            sheet (xlrd.sheet.Sheet): The worksheet object.
        Returns:
            sheet_content (str): The markdown table as a string.
        """
        n_rows = sheet.nrows
        n_cols = sheet.ncols
        first_pct = 10 # max(1, int(n_rows * 0.05))

        # Read first 5% rows
        candidate_rows = []
        for i in range(first_pct):
            row_values = [
                (
                    str(sheet.cell_value(i, j)).replace("\n", "<br>")
                    if sheet.cell_value(i, j) not in [None, ""]
                    else ""
                )
                for j in range(n_cols)
            ]
            candidate_rows.append(row_values)

        # Find header row: most non-empty cells
        filled_counts = [
            sum(1 for cell in row if cell.strip()) for row in candidate_rows
        ]
        header_idx = filled_counts.index(max(filled_counts))

        header_row = candidate_rows[header_idx]
        # Free text rows above header
        markdown_lines = []
        for row in candidate_rows[:header_idx]:
            free_text = " ".join(cell for cell in row if cell.strip())
            if free_text:
                markdown_lines.append(f"{free_text}\n")

        # Markdown table header
        markdown_lines.append("| " + " | ".join(header_row) + " |")
        markdown_lines.append("|" + " --- |" * len(header_row))

        # Read rest of the sheet
        for i in range(header_idx + 1, n_rows):
            row = [
                (
                    str(sheet.cell_value(i, j))
                    if sheet.cell_value(i, j) not in [None, ""]
                    else ""
                )
                for j in range(n_cols)
            ]
            markdown_lines.append("| " + " | ".join(row) + " |")

        sheet_content = "\n".join(markdown_lines)
        return sheet_content

    @async_timer
    async def extract_common_xls(self, sheet) -> str:
        """
        Converts the xlrd sheet into markdown data (LLM block detection).
        Args:
            sheet (xlrd.sheet.Sheet): The worksheet object.
        Returns:
            sheet_content (str): The markdown data.
        """
        n_rows = sheet.nrows
        n_cols = sheet.ncols

        # get the list of cell values with location
        cells_data = []
        trunc_cells = []
        for row_idx in range(n_rows):
            for col_idx in range(n_cols):
                value = sheet.cell_value(row_idx, col_idx)
                if value not in [None, ""]:
                    # xlrd is 0-indexed, so add 1 for row and col
                    col_letter = get_column_letter(col_idx + 1)
                    value_str = self.process_cell_value(value)
                    cells_data.append(
                        [f"{row_idx + 1}{col_letter}", value_str])
                    trunc_cells.append(
                        [f"{row_idx + 1}{col_letter}", value_str[:5]])

        # Step 1: Call LLM to get block info
        block_info = await self.llm_client.detect_block(trunc_cells)

        # Step 2: Process blocks
        markdown_chunks = []
        for block in block_info:
            if block[0] == "text":
                loc = block[1]
                text_cell = next((v for l, v in cells_data if l == loc), "")
                if text_cell:
                    markdown_chunks.append(f"{text_cell}\n")
            elif block[0] == "table":
                start_loc, end_loc = block[1], block[2]
                start_row = int("".join(filter(str.isdigit, start_loc)))
                start_col = "".join(filter(str.isalpha, start_loc))
                end_row = int("".join(filter(str.isdigit, end_loc)))
                end_col = "".join(filter(str.isalpha, end_loc))

                # Convert column letter to index (xlrd is 0-indexed)
                start_col_idx = column_index_from_string(start_col) - 1
                end_col_idx = column_index_from_string(end_col) - 1

                # Extract table block
                table_rows = []
                for r in range(start_row - 1, end_row):  # xlrd rows are 0-indexed
                    row_cells = []
                    for c in range(start_col_idx, end_col_idx + 1):
                        value = sheet.cell_value(r, c)
                        row_cells.append(self.process_cell_value(value) if value not in [
                                         None, ""] else "")
                    table_rows.append(row_cells)

                # Markdown table
                if table_rows:
                    header = table_rows[0]
                    markdown_chunks.append("| " + " | ".join(header) + " |")
                    markdown_chunks.append("|" + " --- |" * len(header))
                    for row in table_rows[1:]:
                        markdown_chunks.append("| " + " | ".join(row) + " |")

        sheet_content = "\n".join(markdown_chunks)
        return sheet_content

    @async_timer
    async def extract_xlsb_markdown(
        self, file_path: Path | str, limit: Tuple = (500, 250)
    ) -> str:
        """
        Extract xlsb file into markdown contents using direct cell access.
        Args:
            file_path (str): path to excel files.
            limit (Tuple): limit max row / max column
        Returns:
            contents (str): markdown content of the xlsb files.
        """
        file_name = Path(file_path).name
        markdown_chunks = [f"# Document: {file_name}"]

        with open_workbook(str(file_path)) as wb:
            for sheet_name in wb.sheets:
                with wb.get_sheet(sheet_name) as sheet:
                    rows = list(sheet.rows())
                    n_rows = len(rows)
                    n_cols = max(len(row) for row in rows) if rows else 0

                    # You may want to write direct-access versions of these:
                    if n_rows >= limit[0] or n_cols >= limit[1]:
                        sheet_content = await self.extract_large_xlsb(rows)
                    else:
                        sheet_content = await self.extract_common_xlsb(rows)

                    markdown_chunks.append(f"\n## Sheet: {sheet_name}\n")
                    markdown_chunks.append(sheet_content)

        markdown_content = "\n".join(markdown_chunks)
        return markdown_content

    @async_timer
    async def extract_large_xlsb(self, rows) -> str:
        """
        Converts the entire .xlsb sheet to a markdown table, auto-detects header.
        Args:
            rows (list of row lists): Each row is a list of pyxlsb cell objects.
        Returns:
            sheet_content (str): The markdown table as a string.
        """
        n_rows = len(rows)
        n_cols = max(len(row) for row in rows) if rows else 0
        first_pct = 10 # max(1, int(n_rows * 0.05))

        # Read first 5% rows
        candidate_rows = []
        for i in range(first_pct):
            row_values = (
                [
                    str(row[j].v).replace("\n", "<br>") if j < len(
                        row) and row[j].v is not None else ""
                    for j in range(n_cols)
                ]
                if i < len(rows)
                else []
            )
            candidate_rows.append(row_values)

        # Find header row: most non-empty cells
        filled_counts = [
            sum(1 for cell in row if cell.strip()) for row in candidate_rows
        ]
        header_idx = filled_counts.index(max(filled_counts))

        header_row = candidate_rows[header_idx]
        # Free text rows above header
        markdown_lines = []
        for row in candidate_rows[:header_idx]:
            free_text = " ".join(cell for cell in row if cell.strip())
            if free_text:
                markdown_lines.append(f"{free_text}\n")

        # Markdown table header
        markdown_lines.append("| " + " | ".join(header_row) + " |")
        markdown_lines.append("|" + " --- |" * len(header_row))

        # Read rest of the sheet
        for i in range(header_idx + 1, n_rows):
            row = [
                (
                    str(rows[i][j].v)
                    if j < len(rows[i]) and rows[i][j].v is not None
                    else ""
                )
                for j in range(n_cols)
            ]
            markdown_lines.append("| " + " | ".join(row) + " |")

        sheet_content = "\n".join(markdown_lines)
        return sheet_content

    @async_timer
    async def extract_common_xlsb(self, rows) -> str:
        """
        Converts the .xlsb sheet into markdown using LLM block detection.
        Args:
            rows (list of row lists): Each row is a list of pyxlsb cell objects.
        Returns:
            sheet_content (str): The markdown data.
        """
        n_rows = len(rows)
        n_cols = max(len(row) for row in rows) if rows else 0

        # Gather cell data
        cells_data = []
        trunc_cells = []
        for row_idx, row in enumerate(rows):
            for col_idx in range(n_cols):
                value = row[col_idx].v if col_idx < len(row) else None
                if value is not None and str(value).strip():
                    col_letter = get_column_letter(col_idx + 1)
                    value_str = self.process_cell_value(value)
                    cells_data.append(
                        [f"{row_idx + 1}{col_letter}", value_str])
                    trunc_cells.append(
                        [f"{row_idx + 1}{col_letter}", value_str[:5]])

        # Step 1: Call LLM to get block info
        block_info = await self.llm_client.detect_block(trunc_cells)

        # Step 2: Process blocks
        markdown_chunks = []
        for block in block_info:
            if block[0] == "text":
                loc = block[1]
                text_cell = next((v for l, v in cells_data if l == loc), "")
                if text_cell:
                    markdown_chunks.append(f"{text_cell}\n")
            elif block[0] == "table":
                start_loc, end_loc = block[1], block[2]
                start_row = int("".join(filter(str.isdigit, start_loc)))
                start_col = "".join(filter(str.isalpha, start_loc))
                end_row = int("".join(filter(str.isdigit, end_loc)))
                end_col = "".join(filter(str.isalpha, end_loc))

                start_col_idx = column_index_from_string(start_col) - 1
                end_col_idx = column_index_from_string(end_col) - 1

                table_rows = []
                for r in range(start_row - 1, end_row):  # pyxlsb is 0-indexed
                    row = rows[r] if r < len(rows) else []
                    row_cells = [
                        # str(row[c].v) if c < len(
                        #     row) and row[c].v is not None else ""
                        # for c in range(start_col_idx, end_col_idx + 1)
                        self.process_cell_value(row[c].v) if c < len(
                            row) and row[c].v is not None else ""
                        for c in range(start_col_idx, end_col_idx + 1)
                    ]
                    table_rows.append(row_cells)

                # Markdown table
                if table_rows:
                    header = table_rows[0]
                    markdown_chunks.append("| " + " | ".join(header) + " |")
                    markdown_chunks.append("|" + " --- |" * len(header))
                    for row in table_rows[1:]:
                        markdown_chunks.append("| " + " | ".join(row) + " |")

        sheet_content = "\n".join(markdown_chunks)
        return sheet_content
